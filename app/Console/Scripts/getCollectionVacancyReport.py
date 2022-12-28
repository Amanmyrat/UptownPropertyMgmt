from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font,PatternFill,Alignment,Border, Side
import sys
import json

excelCollectionFileName = sys.argv[1] 
headersFileName = sys.argv[2]
vacancyCollectionFileName = sys.argv[3] 
vacancyHeadersFileName = sys.argv[4]
saveFilePath = sys.argv[5]

excelCollectionFile = open(excelCollectionFileName,encoding="utf-8",errors='ignore')
headersFile = open(headersFileName,encoding="utf-8",errors='ignore')
vacancyCollectionFile = open(vacancyCollectionFileName,encoding="utf-8",errors='ignore')
vacancyHeadersFile = open(vacancyHeadersFileName,encoding="utf-8",errors='ignore')

excelCollection = json.loads(excelCollectionFile.read())
headers = json.loads(headersFile.read())
vacancyCollection = json.loads(vacancyCollectionFile.read())
vacancyHeaders = json.loads(vacancyHeadersFile.read())
# print(vacancyHeaders)

# Create a Workbook
wb = Workbook()

def createReport(title, data, dataHeaders, start_row, report_name, collectionType = 'collection'):
    ws =  wb.active
    ws.title = title

    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

    cellref = ws.cell(row=start_row, column=1)
    cellref.value = report_name
    cellref.alignment = Alignment(horizontal='center', vertical='center')
    cellref.border = thin_border
    cellref.font = Font(color='F90000', bold=True, size = "20")
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=2)

    for i in range(0,len(dataHeaders)):
        cellref = ws.cell(row=start_row + 1, column=i+1)
        cellref.value= dataHeaders[i]
        cellref.font = Font(bold=True)
        cellref.alignment = Alignment(horizontal='center', vertical='center')
        cellref.border = thin_border
        if(i==1):
            cellref.fill = PatternFill(fgColor="002060", fill_type = "solid")
            cellref.font = Font(color='ffffff', bold=True)

        if(collectionType == 'vacancy'):
            if(dataHeaders[i] == "Vacant Done"):
                cellref.fill = PatternFill(fgColor="00B050", fill_type = "solid")
            elif(dataHeaders[i] == "Vacant Not Done"):
                cellref.fill = PatternFill(fgColor="FF0000", fill_type = "solid")
            elif(dataHeaders[i] == "Down Units"):
                cellref.fill = PatternFill(fgColor="996633", fill_type = "solid")
            elif(dataHeaders[i] == "Burn Units"):
                cellref.fill = PatternFill(fgColor="000000", fill_type = "solid")
                cellref.font = Font(color='FF0000',bold=True)

    for i in range(0,len(data)):
            for j in range(0,len(dataHeaders)):
                cellref = ws.cell(row=i+2+start_row, column=j+1)
                cellref.alignment = Alignment(horizontal='center')
                cellref.border = thin_border
            
                try:
                    val = data[i][j]
                    try:
                        val = float(val)
                    except ValueError:
                        pass
                    
                    cellref.value = val
                    
                    if (i % 2) == 0:
                        if(j==1):
                            cellref.fill = PatternFill(fgColor="66CCFF", fill_type = "solid")
                        elif(j!=0):
                            cellref.fill = PatternFill(fgColor="ffe699", fill_type = "solid")
                            
                    else:
                        if(j==1):
                            cellref.fill = PatternFill(fgColor="00B0F0", fill_type = "solid")
                        elif(j!=0):
                            cellref.fill = PatternFill(fgColor="fff702", fill_type = "solid")
                    
                    if(collectionType == 'collection'):
                        if('Total' in dataHeaders[j].split() or 'Prev Month Today' in dataHeaders[j] or 
                            '12 Months Max Collection' in dataHeaders[j] or 'Max Collection 5% Increase' in dataHeaders[j]):
                            cellref.font = Font(bold=True)
                            cellref.fill = PatternFill(fgColor="B4C6E7", fill_type = "solid")
                        
                            if('TOTAL' in data[i][1].split() or 'total' in data[i][1].split() or 'Total' in data[i][1].split()):
                                cellref.fill = PatternFill(fgColor="33CC33", fill_type = "solid")
                                cellref.font = Font(bold=True)
                                if('Total' in dataHeaders[j].split()):
                                    cellref.font = Font(bold=True)
                                    cellref.fill = PatternFill(fgColor="FFA500", fill_type = "solid")
                                
                                if('Total SQ Feet' in dataHeaders[j] or
                                    'Total Charged Rent To Current Tenant' in dataHeaders[j] or
                                    'Total Uncollected Rent' in dataHeaders[j]):
                                    
                                    cellref.font = Font(bold=True)
                                    cellref.fill = PatternFill(fgColor="32C732", fill_type = "solid")
                    else:
                        if('TOTAL' in data[i][1].split() or 'total' in data[i][1].split() or 'Total' in data[i][1].split()):
                            cellref.fill = PatternFill(fgColor="33CC33", fill_type = "solid")
                            cellref.font = Font(bold=True)
                    if(j==0):
                        cellref.font = Font(bold=True)
                        cellref.fill = PatternFill(fgColor="ffffff", fill_type = "solid")
                                    
                except IndexError:
                    pass
                continue

    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value+3

    ws.row_dimensions[2].height = 30
    ws.freeze_panes = 'C1'


createReport("Excel Collection", excelCollection, headers, 2, "COLLECTION REPORT")
createReport("Excel Collection", vacancyCollection, vacancyHeaders, 8, "VACANCY REPORT", 'vacancy')
wb.save(saveFilePath)
wb.close()  

print('success')
print(saveFilePath)